# Import dependencies
import streamlit as st
import pandas as pd

from openpyxl import load_workbook
from Utilities.create_properties import create_properties
from Utilities.create_workbook import create_workbook
from Utilities.upload_prop_uses import upload_prop_uses
from Utilities.drop_empty_cols import drop_empty_cols

# Set a title for the page
st.markdown("<h1 style = 'text-align: center; color: green;'>Green EconoME</h1>", unsafe_allow_html = True)
st.markdown("<h2 style = 'text-align: center; color: black;'>Property Uploader</h2>", unsafe_allow_html = True)

# Set the domain and headers for the API calls
domain = 'https://portfoliomanager.energystar.gov/ws'
headers = {'Content-Type': 'application/xml'}

# Load credentials for the API calls into the auth variable
credential_upload = st.file_uploader('Upload ESPM API Credentials')
if credential_upload:
    creds = []
    for line in credential_upload:
        creds.append(line.decode().strip())
    auth = (creds[0], creds[1])

st.caption('Upload the API credentials within a .txt file with the Username and Password on seperate lines. <br>' + 
            'The .txt file should be of the following format:<br>' + 
            'Username<br>' + 
            'Password', unsafe_allow_html = True)

# Add a file uploader to upload the building survey
building_survey = st.file_uploader('Upload the Building Survey Spreadsheet')

# Create a boolean to use to only call the download when the properties have finished uploading
complete_upload = False

# Check to make sure the required fields are populated
if credential_upload and building_survey is not None:
    if st.button('Upload Properties'):
        with st.spinner('Uploading Properties'):
            
            # Read the workbook
            workbook = load_workbook(building_survey, read_only = True)
            
            # Create a dictionary to hold the dataframes that have been created
            dfs = {}
            # Read in the worksheets containing the order form and the propery use details
            if 'Order Form' in workbook.sheetnames:
                # Read in the order form
                order_form = pd.read_excel(building_survey, 
                                            sheet_name = 'Order Form', 
                                            skiprows = 7)
                # Drop the empty columns from the order_form df
                order_form = drop_empty_cols(order_form)
                # Create the properties on the order form
                created_props, props_failed_to_create = create_properties(order_form, domain, headers, auth)
                dfs['Created Props'] = created_props
                dfs['Props Failed to Create'] = props_failed_to_create
                st.write('Created Properties')
                st.write(created_props)
                
                if 'Building Details' in workbook.sheetnames:
                    building_details = pd.read_excel(building_survey, 
                                                sheet_name = 'Building Details', 
                                                skiprows = 7)
                    
                    # Drop the empty columns from the building_details df
                    building_details = drop_empty_cols(building_details)

                    # Merge the created properties with the building_details
                    building_details = building_details.merge(created_props, on = 'Property Name')
                    
                    # Drop the rows that do not have an ESPM ID to be used to push the property use types in the building_details df
                    building_details.dropna(subset = 'ESPM ID')
                    st.write('Building Details after merging created dfs and dropping NAs in ESPM ID col')
                    st.dataframe(building_details.astype(str))
                    # Push the property uses for the properties within the building details df
                    prop_uses_failed, successful_prop_use_population = upload_prop_uses(building_details, domain, headers, auth)
                    dfs['Prop Uses Failed to Populate'] = prop_uses_failed
                    dfs['Successful Prop Use Uploads'] = successful_prop_use_population

            # # Merge the two worksheets together
            # property_df = order_form.merge(building_details, on = 'Property Name')

            # Drop the empty columns (inserted into the building survey for formatting)


            # # Call the create_properties function to create each building and populate the property uses
            # (created_props, props_failed_to_create, 
            # props_failed_w_prop_uses, successful_uploads) = create_properties(order_form, domain, headers, auth)

            # Create a workbook containing the successful/unsuccessful uploads
            workbook = create_workbook(dfs)

            # Update the complete_upload
            complete_upload = True

    # If the properties have finished uploading
    if complete_upload:
        
        st.download_button(label = 'Download Property Upload Results', 
                            data = workbook, 
                            file_name = 'Property Upload Results.xlsx')
        

